import openpyxl
wb = openpyxl.load_workbook("example.xlsx")

# List of all the Sheet Names
print(wb.sheetnames)

# Store a sheet as a sheet object
sheet3 = wb["Sheet3"]
print(sheet3.title)

# Store the active worksheet
activesheet = wb.active
print(activesheet.title)

# Get info from a cell A1
print(activesheet['A1'].value)

# Attributes of cells
c2 = activesheet['C2']
print('Row {}, column {}, aka cell {} contains the value: {}\n'.format(c2.row, c2.column, c2.coordinate, c2.value))

# Print the whole excel table

for row in range(1, activesheet.max_row+1):
    for col in range(1, activesheet.max_column+1):
        print(activesheet.cell(row=row, column=col).value, end='\t')
    print()

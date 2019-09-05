"""Program to update the prices in an excel file."""

import openpyxl
from openpyxl.styles import Font

print('Opening workbook...')
wb = openpyxl.load_workbook('xl_produceSales.xlsx')
sheet = wb['Sheet']

# All the prices that need to be updated in xl_produceSales.xlsx
price_updates = {'Garlic': 3.77,
                 'Celery': 1.77,
                 'Lemon': 1.42}

# Font style to set for updated rows
font_style = Font(sz=12,
                  b=True,
                  i=True,
                  color='00FF0000')

# Iterate through all the rows of the excel file
for row in range(2, sheet.max_row+1):
    produce_name = sheet.cell(row=row, column=1).value
    if produce_name in price_updates:
        sheet.cell(row=row, column=2).value = price_updates[produce_name]

        # Update the style of the changed row
        sheet.cell(row=row, column=1).font = font_style

# Freeze panes at cell A2
sheet.freeze_panes = 'A2'

print('Saving new workbook with updated prices...')
wb.save('xl_produceSales_updated.xlsx')
